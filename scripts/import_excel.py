#!/usr/bin/env python3
"""
Import a research Excel sheet into src/data/companies.json.

Merges by company (TradingView code / name): each run upserts the company's
fixed info and adds/replaces ONE quarter entry in its timeline, so running it
every quarter builds up the history instead of overwriting it.

NOTE: the upsert rules here mirror functions/_lib/merge.js (the /admin
upload path). Two differences are intentional: this script reads by column
*index* (the raw source sheet), while merge.mjs reads by header *name* (the
template); and merge.mjs validates REQUIRED_COLUMNS while this local script does
not. Keep the upsert/keying/sort behaviour in sync between the two.

Usage:
    python3 scripts/import_excel.py \
        --input "/path/to/Research Stock 2026.xlsx" \
        --sheet "Q426 Earning" \
        --quarter "Q4 2026"

Requires: openpyxl  (pip install openpyxl)
"""
import argparse, json, math, os, re, sys

try:
    import openpyxl
except ImportError:
    sys.exit("Please install openpyxl:  pip install openpyxl")

# Read by HEADER NAME, not column position. The source sheet's layout has already
# changed once (Industry moved to column 1, the growth columns landed in 3-10), and
# index-based reads fail silently by pulling every field from the wrong place.
# Aliases let one mapping cover both the raw research sheet and the /admin template.
COL_ALIASES = {
    "name":      ["Company Name", "CompanyName"],
    "industry":  ["Industry"],
    "marketCap": ["Market Cap", "MarketCap"],
    "tier":      ["Tier"],
    "tvCode":    ["TradingView Code", "TradingViewCode"],
    "view":      ["View"],
    "note":      ["Note"],
}
# YoY/QoQ growth, percentages. A metric with both cells blank is simply omitted,
# so the site shows "not recorded" rather than a misleading zero.
GROWTH_ALIASES = {
    "sales":    {"yoy": ["YoY Sales Growth", "SalesYoY"],
                 "qoq": ["QoQ Sales Growth", "SalesQoQ"]},
    "opProfit": {"yoy": ["YoY Op Profit Growth", "OpProfitYoY"],
                 "qoq": ["QoQ Op Profit Growth", "OpProfitQoQ"]},
    "eps":      {"yoy": ["YoY EPS Growth", "EPSYoY"],
                 "qoq": ["QoQ EPS Growth", "EPSQoQ"]},
    "pat":      {"yoy": ["YoY PAT Growth", "PATYoY"],
                 "qoq": ["QoQ PAT Growth", "PATQoQ"]},
}
REQUIRED = ["name", "view", "note"]

KNOWN_VIEWS = {"positive", "watch", "concern", "negative"}


def clean_tv(v):
    """A TradingView symbol is short and space-free. The Q1 2027 sheet had a whole
    sentence pasted into that cell, which would have slugged into a 160-character
    id and created a duplicate company — so reject anything that cannot be a
    ticker and fall back to matching on the name."""
    t = str(v).strip() if v is not None else ""
    if not t or " " in t or len(t) > 20:
        return None, (t or None)
    return t, None


def slugify(s):
    return re.sub(r"[^a-z0-9]+", "-", str(s).lower()).strip("-") or "x"


def round2(x):
    # Round half UP to match merge.mjs (JS Math.round) so CLI and /admin agree.
    return math.floor(x * 100 + 0.5) / 100


def parse_mcap(v):
    # Accept numeric-looking text ("2,458.10", "₹1,234") too, else it drops to None.
    if isinstance(v, (int, float)):
        return round2(float(v))
    if v is None:
        return None
    s = re.sub(r"[^0-9.\-]", "", str(v))
    try:
        return round2(float(s)) if s not in ("", "-", ".") else None
    except ValueError:
        return None


def build_header_map(ws):
    """{normalised header -> column index} from row 1."""
    out = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v is None:
            continue
        out[re.sub(r"[^a-z0-9]", "", str(v).lower())] = c
    return out


def resolve(hmap, aliases):
    for a in aliases:
        c = hmap.get(re.sub(r"[^a-z0-9]", "", a.lower()))
        if c:
            return c
    return None


def cell(ws, r, col):
    return ws.cell(row=r, column=col).value if col else None


def parse_pct(v):
    """'24.1', '24.1%', 24.1 -> 24.1 ; blank/garbage -> None (never 0)."""
    if isinstance(v, (int, float)):
        return round2(float(v))
    if v is None:
        return None
    s = re.sub(r"[^0-9.\-]", "", str(v))
    if s in ("", "-", ".", "-."):
        return None
    try:
        return round2(float(s))
    except ValueError:
        return None


def read_growth(ws, r, gcols):
    """Nested under one key so a quarter with no figures just omits it — every
    existing entry in companies.json stays valid, so there is no migration."""
    out = {}
    for metric, cols in gcols.items():
        m = {}
        for basis in ("yoy", "qoq"):
            v = parse_pct(cell(ws, r, cols.get(basis)))
            if v is not None:
                m[basis] = v
        if m:
            out[metric] = m
    return out or None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", required=True, help="Path to the .xlsx file")
    ap.add_argument("--sheet", default="Q426 Earning", help="Worksheet name")
    ap.add_argument("--quarter", required=True, help='Quarter label, e.g. "Q4 2026"')
    here = os.path.dirname(os.path.abspath(__file__))
    ap.add_argument("--json", default=os.path.join(here, "..", "src", "data", "companies.json"))
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.input, data_only=True)
    try:
        ws = wb[args.sheet]
    except KeyError:
        sys.exit(f"Sheet '{args.sheet}' not found. Available sheets: {', '.join(wb.sheetnames)}")

    # Resolve every field from the header row, then prove the mapping is real by
    # checking the View column actually holds views. A missing header is a loud
    # failure here rather than a silent wrong-column read later.
    hmap = build_header_map(ws)
    cols = {k: resolve(hmap, a) for k, a in COL_ALIASES.items()}
    missing = [k for k in REQUIRED if not cols.get(k)]
    if missing:
        sys.exit(f"Sheet '{args.sheet}': no column found for {missing}. "
                 f"Headers seen: {', '.join(str(ws.cell(row=1, column=c).value) for c in range(1, ws.max_column + 1) if ws.cell(row=1, column=c).value)}")
    gcols = {m: {b: resolve(hmap, a) for b, a in bases.items()} for m, bases in GROWTH_ALIASES.items()}
    found_growth = sorted(m for m, b in gcols.items() if any(b.values()))
    seen_views = sum(
        1 for r in range(2, ws.max_row + 1)
        if str(cell(ws, r, cols["view"]) or "").strip().lower() in KNOWN_VIEWS
    )
    if ws.max_row > 5 and seen_views == 0:
        sys.exit("No recognizable View values (Positive/Watch/Concern/Negative) in the "
                 "resolved View column — the sheet layout looks wrong. Aborting.")
    print(f"Columns resolved from headers. Growth metrics found: "
          f"{', '.join(found_growth) if found_growth else 'none'}")

    # load existing dataset (keeps prior quarters)
    target = os.path.abspath(args.json)
    if os.path.exists(target):
        with open(target) as f:
            data = json.load(f)
    else:
        data = {"companies": []}

    def nkey(t):
        return re.sub(r"[^a-z0-9]", "", str(t or "").lower())

    by_key, by_name = {}, {}
    for c in data["companies"]:
        by_key[str(c.get("tvCode") or c.get("name") or "").upper()] = c
        by_name[nkey(c.get("name"))] = c
    bad_tv = []

    added, updated = 0, 0
    for r in range(2, ws.max_row + 1):
        name = cell(ws, r, cols["name"])
        if not name:
            continue
        tv, tv_junk = clean_tv(cell(ws, r, cols["tvCode"]))
        if tv_junk:
            bad_tv.append((r, str(name).strip(), tv_junk))
        key = (str(tv) if tv else str(name)).upper()

        view = cell(ws, r, cols["view"])
        note = cell(ws, r, cols["note"])
        tier = cell(ws, r, cols["tier"])
        mcap = cell(ws, r, cols["marketCap"])
        industry = cell(ws, r, cols["industry"])

        mcap_val = parse_mcap(mcap)
        comp = by_key.get(key) or by_name.get(nkey(name))
        is_new = comp is None
        if comp is None:
            comp = {
                "slug": slugify(tv or name),
                "name": str(name).strip(),
                "tvCode": (str(tv).strip() if tv else None),
                "industry": (str(industry).strip() if industry else None),
                "marketCap": mcap_val,
                "tier": (str(tier).strip() if tier else None),
                "quarters": [],
            }
            data["companies"].append(comp)
            by_key[key] = comp
            by_name[nkey(name)] = comp
            by_key[key] = comp
        else:
            # refresh latest fixed info
            if mcap_val is not None:
                comp["marketCap"] = mcap_val
            if industry:
                comp["industry"] = str(industry).strip()
            if tier:
                comp["tier"] = str(tier).strip()

        # Count only rows that produce a quarter entry — others are dropped below.
        if (view and str(view).strip()) or (note and str(note).strip()):
            prior = next((q for q in comp["quarters"] if q.get("quarter") == args.quarter), None)
            entry = {
                "quarter": args.quarter,
                "tier": (str(tier).strip() if tier else None),
                "view": (str(view).strip() if view else None),
                "note": (str(note).strip() if note else ""),
            }
            growth = read_growth(ws, r, gcols)
            if growth:
                entry["growth"] = growth
            # The CLI path never re-runs linkReports, so keep any existing link.
            if prior and prior.get("reportUrl"):
                entry["reportUrl"] = prior["reportUrl"]
            comp["quarters"] = [q for q in comp["quarters"] if q.get("quarter") != args.quarter]
            comp["quarters"].append(entry)
            if is_new:
                added += 1
            else:
                updated += 1

    # keep only companies that have at least one quarter's research note
    data["companies"] = [c for c in data["companies"] if c.get("quarters")]

    # market-cap desc, unknown last
    data["companies"].sort(key=lambda c: (c.get("marketCap") is None, -(c.get("marketCap") or 0)))

    os.makedirs(os.path.dirname(target), exist_ok=True)
    with open(target, "w") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    for r, nm, junk in bad_tv:
        print(f"  ! row {r} {nm}: TradingView Code cell is not a symbol "
              f"({junk[:60]!r}) — matched by name instead; fix the sheet.")
    print(f"{args.quarter}: {added} new, {updated} existing companies. "
          f"Total: {len(data['companies'])} -> {target}")


if __name__ == "__main__":
    main()
